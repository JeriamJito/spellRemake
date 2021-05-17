from PyQt6 import uic, QtWidgets
from openpyxl import *
from random import randint, choice
from chronometer import *
from math import trunc
from operator import itemgetter


def pontuacao():
    global sortedranking
    pontos = janelaPrincipal.barraPontos.value()
    nome = janelanome.NomeDigitado.text()
    wb = load_workbook("placar.xlsx")
    ws = wb.active
    rankingtotal = dict()
    rankingtotal[nome] = pontos
    valatual = ws['C2'].value
    sortedranking = list()

    if valatual == 12:
        # checar se tem algum nome igual na lista e se quebrou o recorde
        for i in range(1, valatual + 1):
            if nome.lower() == ws[f"A{i}"].value.lower() and ws[f"B{i}"].value < pontos:
                ws[f"B{i}"] = pontos

        # coloca itens da planilha em um dicionario
        for n in range(1, valatual + 1):
            rankingtotal[f"{ws[f'A{n}'].value}"] = ws[f"B{n}"].value
        # organiza os itens do dicionario numa lista
        sortedranking = sorted(rankingtotal.items(), key=itemgetter(1), reverse=True)

        # insere itens da lista organizada na tabela do XL
        for i in range(1, valatual + 1):
            imenosum = i - 1
            ws[f"A{i}"] = sortedranking[imenosum][0]
            ws[f"B{i}"] = sortedranking[imenosum][1]

    wb.save("placar.xlsx")
    janelanome.close()
    janelaplacar.show()


def mostranomes():
    global sortedranking

    nomesnoplacar = [
        janelaplacar.nome1, janelaplacar.nome2, janelaplacar.nome3, janelaplacar.nome4, janelaplacar.nome5,
        janelaplacar.nome6, janelaplacar.nome7, janelaplacar.nome8, janelaplacar.nome9, janelaplacar.nome10,
        janelaplacar.nome11, janelaplacar.nome12
    ]
    pontosnoplacar = [
        janelaplacar.ponto1, janelaplacar.ponto2, janelaplacar.ponto3, janelaplacar.ponto4, janelaplacar.ponto5,
        janelaplacar.ponto6, janelaplacar.ponto7, janelaplacar.ponto8, janelaplacar.ponto9, janelaplacar.ponto10,
        janelaplacar.ponto11, janelaplacar.ponto12
    ]
    for i in range(0, 12):
        nomesnoplacar[i].setText(sortedranking[i][0])
        pontosnoplacar[i].setText(str(sortedranking[i][1]))


def comecaJogo():
    global n
    global pal_port
    global dica_port
    global pal_ing
    global dica_ing
    global temporestante
    global tempoMax
    if len(selecao) == 0:
        janela_msgCertoeErrado.certoerrado.setText("Error!")
        janela_msgCertoeErrado.frasedeefeito.setText("you didn't selected any filter for the game or you selected an empty filter.")
        janela_msgCertoeErrado.botaoConfirma.clicked.connect(janela_msgCertoeErrado.close)
        janela_msgCertoeErrado.show()
    else:
        janelaPrincipal.barraPontos.setValue(5)

        janelaPrincipal.valAcertos.setText(str(janelaPrincipal.barraAcertos.value()))
        janelaPrincipal.valPontos.setText(str(janelaPrincipal.barraPontos.value()))
        janelaPrincipal.valErros.setText(str(janelaPrincipal.barraErros.value()))
        tempoMax = len(selecao) * 5
        # manda mudar o tempo
        mudatempo()
        # faz uso da classe cronômetro
        cronometro = Chronometer()

        n = choice(selecao)
        pal_port = ws[f"A{n}"].value
        dica_port = ws[f"B{n}"].value
        pal_ing = ws[f"C{n}"].value
        dica_ing = ws[f"D{n}"].value
        temporestante = tempoMax - trunc(float(cronometro))
        janelaPrincipal.barraTempo.setValue(temporestante)

        # Comando para iniciar o cronômetro
        cronometro.start()

        mudaTexto(pal_port, dica_ing)
        janelaPrincipal.show()
        janelafiltro.close()


def teladefiltros():
    principal()
    janelafiltro.BotaoOK.clicked.connect(printa)
    janelafiltro.BotaoOK.clicked.connect(comecaJogo)
    janelafiltro.show()
    menuPrincipal.close()




def mudaTexto(palavratexto, palavraescrita):
    janelaPrincipal.dica_ing.setText(palavraescrita)
    janelaPrincipal.pal_port.setText(palavratexto)


def mudatempo():
    janelaPrincipal.barraTempo.setMaximum(tempoMax)


def confirma():
    global n
    global pal_port
    global dica_port
    global pal_ing
    global dica_ing
    global temporestante
    global barravermelha
    global avermelhou
    acertos = ["Seu Sortudo da desgraça!", "Albert Einstein teria inveja de você neste exato momento.",
               "Isso é sorte de principiante.", "você tá com uma sorte daquelas, viu.",
               "Sua inteligência chega me irrita."]
    erros = ["Seu animal!", "pra ser burro só falta asa!",
             "Não é possível que um monte de pixels numa tela conseguem ser melhor do que você.",
             "ouvi dizer que os cursos de inglês não te aceitam porque você é burro demais",
             "Até meu cachorro faz melhor do que isso."]

    if janelaPrincipal.dica_ing.text() == pal_ing:
        global pontos
        QtWidgets.QMessageBox.about(janelaPrincipal, "Confirmação", "Parabéns, você acertou!")
        janelaPrincipal.barraPontos.setValue(janelaPrincipal.barraPontos.value() + 2)
        janelaPrincipal.barraAcertos.setValue(janelaPrincipal.barraAcertos.value() + 1)
    else:
        QtWidgets.QMessageBox.about(janelaPrincipal, "Pane", f"Você errou. O correto é: {pal_ing}")
        try:
            janelaPrincipal.barraPontos.setValue(janelaPrincipal.barraPontos.value() - 3)
            janelaPrincipal.barraErros.setValue(janelaPrincipal.barraErros.value() + 1)
            if janelaPrincipal.barraPontos.value() <= 2:
                janelaPrincipal.barraPontos.setValue(0)
        except:
            janelaPrincipal.barraPontos.setValue(0)
            janelaPrincipal.barraErros.setValue(janelaPrincipal.barraErros.value() + 1)

    if janelaPrincipal.barraErros.value() % 7 == 0 and janelaPrincipal.barraErros.value() != 1:
        global reperro
        if reperro != janelaPrincipal.barraErros.value():
            escolhe = choice(erros)
            janela_msgCertoeErrado.certoerrado.setText("Error Streak.")
            janela_msgCertoeErrado.frasedeefeito.setText(escolhe)
            reperro = janelaPrincipal.barraErros.value()
            janela_msgCertoeErrado.botaoConfirma.clicked.connect(janela_msgCertoeErrado.close)
            janela_msgCertoeErrado.show()

    if janelaPrincipal.barraAcertos.value() % 7 == 0 and janelaPrincipal.barraAcertos.value() != 1:
        global repacerto
        if repacerto != janelaPrincipal.barraAcertos.value():
            escolhe = choice(acertos)
            janela_msgCertoeErrado.certoerrado.setText("Winning Streak!")
            janela_msgCertoeErrado.frasedeefeito.setText(escolhe) #QtWidgets.QMessageBox.about(janelaPrincipal, "Winning Streak", escolhe)
            repacerto = janelaPrincipal.barraAcertos.value()
            janela_msgCertoeErrado.botaoConfirma.clicked.connect(janela_msgCertoeErrado.close)
            janela_msgCertoeErrado.show()

    if temporestante <= 0:
        janelaPrincipal.barraTempo.setValue(0)
        QtWidgets.QMessageBox.about(janelaPrincipal, "FIM DE JOGO", f"FIM DE JOGO! Sua pontuação foi de: {janelaPrincipal.barraPontos.value()} ponto(s)!")
        janelaPrincipal.close()
        janelanome.NomeDigitado.setText("")
        janelanome.NomeDigitado.returnPressed.connect(pontuacao)
        janelanome.NomeDigitado.returnPressed.connect(mostranomes)
        janelanome.show()



    janelaPrincipal.valAcertos.setText(str(janelaPrincipal.barraAcertos.value()))
    janelaPrincipal.valPontos.setText(str(janelaPrincipal.barraPontos.value()))
    janelaPrincipal.valErros.setText(str(janelaPrincipal.barraErros.value()))


    #crio uma variável local que calcula o tempo restante.
    temporestante = tempoMax - trunc(float(cronometro))
    janelaPrincipal.barraTempo.setValue(temporestante)
    if janelaPrincipal.barraErros.value() == janelaPrincipal.barraErros.maximum():
        janelaPrincipal.barraErros.setMaximum(janelaPrincipal.barraErros.maximum() + 100)
        janelaPrincipal.barraErros.setStyleSheet(barravermelha)

    if janelaPrincipal.barraAcertos.value() == janelaPrincipal.barraAcertos.maximum():
        janelaPrincipal.barraAcertos.setMaximum(janelaPrincipal.barraAcertos.maximum() + 100)
        janelaPrincipal.barraAcertos.setStyleSheet(barravermelha)

    if janelaPrincipal.barraPontos.value() == janelaPrincipal.barraPontos.maximum():
        janelaPrincipal.barraPontos.setMaximum(janelaPrincipal.barraPontos.maximum() + 100)
        janelaPrincipal.barraPontos.setStyleSheet(barravermelha)
        avermelhou = True

    elif janelaPrincipal.barraPontos.value() < janelaPrincipal.barraPontos.maximum() - 1 and avermelhou is True:
        janelaPrincipal.barraPontos.setMaximum(janelaPrincipal.barraPontos.maximum() - 100)
        janelaPrincipal.barraPontos.setStyleSheet("")

    n = randint(2, total)
    pal_port = ws[f"A{n}"].value
    dica_port = ws[f"B{n}"].value
    pal_ing = ws[f"C{n}"].value
    dica_ing = ws[f"D{n}"].value

    mudaTexto(pal_port, dica_ing)


app = QtWidgets.QApplication([])

#lugar de criação das janelas que interpretam a UI
janelaPrincipal = uic.loadUi("telaPrincipal.ui")
menuPrincipal = uic.loadUi("menuPrincipal.ui")
janelanome = uic.loadUi("DigitaNome.ui")
janelaplacar = uic.loadUi("placar.ui")
janela_msgCertoeErrado = uic.loadUi("msg_CertoErrado.ui")

#lugar de criação de componentes necessários pra iniciar o código
barravermelha = """QProgressBar::chunk 
                  {
                    background-color: red;
                  }"""

wb = load_workbook("Spell 2 Database.xlsx")
ws = wb.active
avermelhou = False
total = 593
jogada = 0
pontos = 0
tempoMax = 30
#manda mudar o tempo
mudatempo()
#faz uso da classe cronômetro
cronometro = Chronometer()

n = randint(2, total)
pal_port = ws[f"A{n}"].value
dica_port = ws[f"B{n}"].value
pal_ing = ws[f"C{n}"].value
dica_ing = ws[f"D{n}"].value
temporestante = tempoMax - trunc(float(cronometro))
janelaPrincipal.barraTempo.setValue(temporestante)

#Comando para iniciar o cronômetro
cronometro.start()

mudaTexto(pal_port, dica_ing)

#Cronometro só vai mudar quando for confirmado o total aqui
janelaPrincipal.botaoConfirma.clicked.connect(confirma)
janelaPrincipal.dica_ing.returnPressed.connect(confirma)

#0------0


def printa():
    global selecao
    if janelafiltro.Cognatas.isChecked():
        selecao += Cognatas

    if janelafiltro.Adjetivo.isChecked():
        selecao += adjetivo

    if janelafiltro.Adverbio.isChecked():
        selecao += adverbio

    if janelafiltro.AreaFiscal.isChecked():
        selecao += Area_fiscal

    if janelafiltro.Corpo.isChecked():
        selecao += Corpo

    if janelafiltro.Expressoes.isChecked():
        selecao += expressoes

    if janelafiltro.FalsoCognato.isChecked():
        selecao += falso_cognata

    if janelafiltro.Girias.isChecked():
        selecao += girias

    if janelafiltro.Infinitivo.isChecked():
        selecao += Infinitivo

    if janelafiltro.Participio.isChecked():
        selecao += participio

    if janelafiltro.Passado.isChecked():
        selecao += passado

    if janelafiltro.Perguntas.isChecked():
        selecao += Perguntas

    if janelafiltro.Preposicao.isChecked():
        selecao += preposicao

    if janelafiltro.Sentencas.isChecked():
        selecao += Sentencas

    if janelafiltro.SubAbstrato.isChecked():
        selecao += subs_abstrato

    if janelafiltro.SubConcreto.isChecked():
        selecao += sub_concreto

    return selecao


wb = load_workbook("Spell 2 Database.xlsx")
ws = wb.active
total = 593
Cognatas = list()
selecao = list()
falso_cognata = list()
sub_concreto = list()
subs_abstrato = list()
adjetivo = list()
Infinitivo = list()
passado = list()
participio = list()
Sentencas = list()
expressoes = list()
girias = list()
Perguntas = list()
preposicao = list()
adverbio = list()
Corpo = list()
Area_fiscal = list()


def principal():
    for i in range(2, total):
        if ws[f"E{i}"].value == "X":
            Cognatas.append(i)

        elif ws[f"F{i}"].value == "X":
            falso_cognata.append(i)

        elif ws[f"G{i}"].value == "X":
            sub_concreto.append(i)

        elif ws[f"H{i}"].value == "X":
            subs_abstrato.append(i)

        elif ws[f"I{i}"].value == "X":
            adjetivo.append(i)

        elif ws[f"J{i}"].value == "X":
            Infinitivo.append(i)

        elif ws[f"K{i}"].value == "X":
            passado.append(i)

        elif ws[f"L{i}"].value == "X":
            participio.append(i)

        elif ws[f"M{i}"].value == "X":
            Sentencas.append(i)

        elif ws[f"N{i}"].value == "X":
            expressoes.append(i)

        elif ws[f"O{i}"].value == "X":
            girias.append(i)

        elif ws[f"P{i}"].value == "X":
            Perguntas.append(i)

        elif ws[f"Q{i}"].value == "X":
            preposicao.append(i)

        elif ws[f"R{i}"].value == "X":
            adverbio.append(i)

        elif ws[f"S{i}"].value == "X":
            Corpo.append(i)

        elif ws[f"T{i}"].value == "X":
            Area_fiscal.append(i)

janelafiltro = uic.loadUi("selecaodefiltro.ui")
#0-----0

reperro = 0
repacerto = 0
sortedranking = list()
menuPrincipal.Jogar.clicked.connect(teladefiltros)
menuPrincipal.show()
app.exec()

#-------#


